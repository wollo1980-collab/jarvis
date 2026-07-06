"""
Excel-Lesen (v0.5 Phase 1, ADR-014): liest .xlsx/.xlsm-Dateien read-only
über openpyxl. Sicherheitsstufe 0 (Handbook Kap. 10, v3.3) - reine
Leseaktion, keine Bestätigung nötig.

Bewusst NICHT enthalten (siehe ADR-013/ADR-014): Schreiben, Formatieren,
Power Query, Makros, .xls (Legacy-Format) sowie eine KI-Zusammenfassung
im Command selbst - Letzteres übernimmt commands/reports.py (ADR-015),
das die Lesefunktion hier wiederverwendet statt sie zu duplizieren.
"""
from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from core.models import Plan, Result, Status

logger = logging.getLogger("jarvis.commands.excel")

_SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}

# Obergrenzen gegen unbegrenzten Speicher-/Ausgabeverbrauch bei grossen
# Dateien (benannte Konstanten statt Magic Values, Handbook Kap. 5).
_MAX_ROWS_PER_SHEET = 500
_PREVIEW_ROWS_IN_MESSAGE = 5


class ExcelReadError(Exception):
    """Erwarteter Lesefehler (Datei fehlt/falsches Format/unbekanntes
    Blatt) - wird von den aufrufenden Commands in ein Result übersetzt.
    Technische openpyxl-Fehler (kaputte Datei etc.) werden NICHT hier,
    sondern vom Aufrufer als generische Exception abgefangen."""


def read_workbook_sheets(
    path: Path, sheet: str | None = None
) -> tuple[dict[str, list[tuple]], list[str]]:
    """Liest eine .xlsx/.xlsm-Datei read-only über openpyxl und gibt
    (sheets_data, summary_parts) zurück - sheets_data pro Blatt auf
    _MAX_ROWS_PER_SHEET begrenzt. Wiederverwendet von ReadExcelCommand
    und commands/reports.py (ADR-015), damit die Lese-Logik nur an
    einer Stelle existiert (Kap. 4, DRY)."""
    if not path.exists():
        raise ExcelReadError(f"Datei nicht gefunden: {path}")

    if path.suffix.lower() not in _SUPPORTED_SUFFIXES:
        raise ExcelReadError(
            f"Ich kann aktuell nur .xlsx/.xlsm lesen - '{path.suffix}' gehört nicht dazu."
        )

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet and sheet not in workbook.sheetnames:
            raise ExcelReadError(
                f"Arbeitsblatt '{sheet}' nicht gefunden. Verfügbar: {', '.join(workbook.sheetnames)}"
            )

        sheet_names = [sheet] if sheet else list(workbook.sheetnames)

        sheets_data: dict[str, list[tuple]] = {}
        summary_parts = []
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            rows = []
            for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                if i >= _MAX_ROWS_PER_SHEET:
                    break
                rows.append(row)
            sheets_data[sheet_name] = rows
            summary_parts.append(
                f"{sheet_name} ({worksheet.max_row} Zeile(n) x {worksheet.max_column} Spalte(n))"
            )
        return sheets_data, summary_parts
    finally:
        # read_only-Workbooks halten sonst einen offenen Dateihandle
        # (unter Windows problematisch, wenn die Datei anschliessend
        # in Excel geoeffnet werden soll).
        workbook.close()


class ReadExcelCommand:
    name = "read_excel"
    # Beschreibung ist bewusst so ausführlich, dass core/ai.py KEINEN
    # eigenen Sonderfall für diesen Intent braucht (anders als bei
    # remember_fact/forget_fact) - die Registry-basierte Prompt-
    # Erzeugung (ADR-007) reicht dafür aus.
    description = (
        "Liest eine Excel-Datei (.xlsx/.xlsm, nur Lesen, Sicherheitsstufe 0). "
        "target = vollständiger Dateipfad (z. B. 'C:\\Reports\\beispiel.xlsx'). "
        "Optional: parameters.sheet für ein bestimmtes Arbeitsblatt, sonst "
        "werden alle Blätter gelesen."
    )
    # Reine Leseaktion (Sicherheitsstufe 0) - keine Bestätigung nötig.
    requires_confirmation = False

    def execute(self, plan: Plan) -> Result:
        path_str = (plan.target or "").strip()
        if not path_str:
            return Result(
                status=Status.NEEDS_CLARIFICATION,
                message="Welche Excel-Datei soll ich lesen?",
            )

        path = Path(path_str)
        try:
            sheets_data, summary_parts = read_workbook_sheets(path, plan.parameters.get("sheet"))
        except ExcelReadError as e:
            return Result(status=Status.FAILED, message=str(e))
        except Exception as e:
            return Result(status=Status.FAILED, message=f"Excel-Datei konnte nicht gelesen werden: {e}")

        message = f"{path.name}: {len(sheets_data)} Arbeitsblatt(e) - " + ", ".join(summary_parts)
        return Result(status=Status.SUCCESS, message=message, data={"sheets": sheets_data})


# Registrierungspunkt für dieses Modul - commands/__init__.py liest
# diese Liste beim Start ein.
COMMANDS = [ReadExcelCommand()]
